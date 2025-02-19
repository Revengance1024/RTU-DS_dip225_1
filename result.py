from openpyxl import load_workbook

# Validate that value is a positive number
def validate_number(value):
    try:
        return 0 < float(value)
    except (TypeError, ValueError):
        return False


wb = load_workbook(filename="test1.xlsx")
ws = wb.active

total = 0
errors = 0

for row in ws.iter_rows(min_row=2):
    total += 1

    if not validate_number(row[1].value) or not validate_number(row[2].value):
        errors += 1
        print("id: {:2.0f}, salary: ERROR".format(row[0].value))
        continue

    salary = row[1].value * row[2].value
    # ws["D" + str(row[0].row)] = salary
    print("id: {:2.0f}, salary: {:7.2f}".format(row[0].value, salary))


print("--- Total: {}, Errors: {} ---".format(total, errors))
wb.close()
